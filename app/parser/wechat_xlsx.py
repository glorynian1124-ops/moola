"""微信账单 XLSX 解析器。

兼容微信官方导出的 .xlsx 账单：
- 前若干行为说明行（标题/昵称/时间范围/统计），中间有一行分隔线，
  之后是表头行（含「交易时间」+「金额(元)」）+ 数据行。
- 列：交易时间, 交易类型, 交易对方, 商品, 收/支, 金额(元), 支付方式, 当前状态, 交易单号, 商户单号, 备注
- 金额为数字；收/支 决定方向（支出取负）；中性交易跳过；支出且状态「已全额退款」跳过（原单作废）。
"""
import datetime
from pathlib import Path

import openpyxl


def parse_wechat_xlsx(path: str | Path) -> list[dict]:
    """解析微信账单 xlsx，返回标准交易记录列表。

    每条记录：{amount, category, merchant, note, trans_time, source, raw_data}
    - amount：支出为负，收入为正
    - category：暂为空，由 classify 步骤填充
    - source：xlsx_wechat
    """
    p = Path(path)
    wb = openpyxl.load_workbook(p, data_only=True, read_only=True)
    try:
        ws = wb[wb.sheetnames[0]]
        rows = list(ws.iter_rows(values_only=True))
    finally:
        wb.close()

    # 动态定位表头行（含「交易时间」和「金额」）
    header_idx = None
    for i, row in enumerate(rows):
        if not row:
            continue
        first = str(row[0] or "")
        if "交易时间" in first and any(c and "金额" in str(c) for c in row):
            header_idx = i
            break
    if header_idx is None:
        raise ValueError(f"未找到微信账单表头：{p}")

    header = rows[header_idx]

    def col(name: str):
        for j, c in enumerate(header):
            if c and name in str(c):
                return j
        return None

    i_time = col("交易时间")
    i_inout = col("收/支")
    i_amount = col("金额")
    i_peer = col("交易对方")
    i_goods = col("商品")
    i_status = col("当前状态")
    i_note = col("备注")
    if None in (i_time, i_inout, i_amount):
        raise ValueError("表头缺少关键列（交易时间/收/支/金额）")

    records = []
    for row in rows[header_idx + 1:]:
        if not row or not row[i_time]:
            continue
        t = row[i_time]
        if not isinstance(t, (datetime.datetime, datetime.date, str)):
            continue

        inout = str(row[i_inout] or "").strip()
        try:
            amount = float(row[i_amount])
        except (TypeError, ValueError):
            continue
        if inout == "支出":
            amount = -abs(amount)
        elif inout == "收入":
            amount = abs(amount)
        else:
            continue  # 中性交易（如零钱存取/理财）跳过

        # 支出原单已全额退款 → 作废（退款金额已作为独立收入记录列出）
        status = str(row[i_status] or "").strip() if i_status is not None else ""
        if inout == "支出" and status == "已全额退款":
            continue

        merchant = str(row[i_peer] or "").strip() if i_peer is not None else ""
        goods = str(row[i_goods] or "").strip() if i_goods is not None else ""
        note = str(row[i_note] or "").strip() if i_note is not None else ""
        full_note = "/".join(x for x in (goods, note) if x)

        if isinstance(t, datetime.datetime):
            trans_time = t.strftime("%Y-%m-%d %H:%M:%S")
        elif isinstance(t, datetime.date):
            trans_time = t.strftime("%Y-%m-%d")
        else:
            trans_time = str(t).strip().replace("/", "-")

        records.append({
            "amount": amount,
            "category": "",
            "merchant": merchant,
            "note": full_note,
            "trans_time": trans_time,
            "source": "xlsx_wechat",
            "raw_data": "|".join("" if c is None else str(c) for c in row),
        })
    return records


if __name__ == "__main__":
    import sys
    if len(sys.argv) < 2:
        print("用法: python3 -m app.parser.wechat_xlsx <微信账单.xlsx>")
        sys.exit(1)
    rs = parse_wechat_xlsx(sys.argv[1])
    print(f"共解析 {len(rs)} 笔交易")
    income = sum(r["amount"] for r in rs if r["amount"] > 0)
    expense = sum(-r["amount"] for r in rs if r["amount"] < 0)
    print(f"收入 {sum(1 for r in rs if r['amount'] > 0)} 笔 {income:.2f} 元；"
          f"支出 {sum(1 for r in rs if r['amount'] < 0)} 笔 {expense:.2f} 元")
    for r in rs[:8]:
        print(r)
